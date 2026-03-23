# -*- coding: utf-8 -*-

# --------------------------------------------------------------------------------
# Descrição: Renomeia colunas de arquivos CSV e Excel de forma reutilizável.
# Exemplo de uso: Esta função é usada pela CLI em src/run.py e por wrappers finos
#                 em src/scripts/.
#
# Autor: Jules
# Criado em: 23/03/2026
# Versão: 1.0
#
# Modificado por: Jules
# Modificado em: 23/03/2026
# Licença: MIT
# --------------------------------------------------------------------------------

import csv
import io
import logging
import os
import shutil
import tempfile
from pathlib import Path
from typing import Any, Dict, Optional, Sequence

from openpyxl import load_workbook

from src.phases.phase01_discovery.file_type_specific.csv.delimiter_detector import (
    detect_csv_delimiter,
)


SUPPORTED_FILE_EXTENSIONS = {".csv", ".xlsx"}


def rename_columns_in_file(
    input_file: str,
    old_columns: Sequence[str],
    new_columns: Sequence[str],
    output_file: Optional[str] = None,
    delimiter: Optional[str] = None,
    sheet_name: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Renomeia colunas de um arquivo de dados com base em listas paralelas.

    Args:
        input_file (str): Caminho para o arquivo de entrada.
        old_columns (Sequence[str]): Lista com os nomes atuais das colunas.
        new_columns (Sequence[str]): Lista com os novos nomes das colunas.
        output_file (Optional[str]): Caminho do arquivo de saída. Quando ausente,
            o arquivo de entrada é sobrescrito.
        delimiter (Optional[str]): Delimitador do CSV. Quando ausente, tenta
            detectar automaticamente.
        sheet_name (Optional[str]): Nome da planilha para arquivos Excel.

    Returns:
        Dict[str, Any]: Dicionário com o status da operação e o caminho de saída.
    """
    logger = logging.getLogger(__name__)

    try:
        input_path = Path(input_file).expanduser().resolve()
        output_path = Path(output_file).expanduser().resolve() if output_file else input_path

        _validate_paths(input_path=input_path, output_path=output_path)
        rename_map = _validate_rename_inputs(old_columns=old_columns, new_columns=new_columns)
        _validate_output_extension(input_path=input_path, output_path=output_path)

        file_extension = input_path.suffix.lower()
        if file_extension == ".csv":
            renamed_count = _rename_csv_columns(
                input_path=input_path,
                output_path=output_path,
                rename_map=rename_map,
                delimiter=delimiter,
            )
        elif file_extension == ".xlsx":
            renamed_count = _rename_xlsx_columns(
                input_path=input_path,
                output_path=output_path,
                rename_map=rename_map,
                sheet_name=sheet_name,
            )
        elif file_extension == ".xls":
            raise ValueError(
                "Arquivos .xls não são suportados para renomeação de colunas. "
                "Converta o arquivo para .xlsx ou .csv antes de executar a operação."
            )
        else:
            raise ValueError(
                f"Extensão de arquivo não suportada para renomeação: {input_path.suffix}."
            )

        message = (
            f"Renomeação de colunas concluída com sucesso. "
            f"{renamed_count} coluna(s) alterada(s)."
        )
        logger.info(message)
        return {
            "status": "success",
            "message": message,
            "report_path": str(output_path),
            "renamed_count": renamed_count,
        }
    except (
        FileNotFoundError,
        IsADirectoryError,
        PermissionError,
        ValueError,
        OSError,
    ) as error:
        logger.error("Erro durante a renomeação de colunas: %s", error)
        return {
            "status": "error",
            "message": str(error),
            "report_path": None,
            "renamed_count": 0,
        }


def _validate_paths(input_path: Path, output_path: Path) -> None:
    """
    Valida os caminhos de entrada e saída da operação.

    Args:
        input_path (Path): Caminho do arquivo de entrada.
        output_path (Path): Caminho do arquivo de saída.

    Returns:
        None: Esta função apenas valida os caminhos recebidos.
    """
    if not input_path.exists():
        raise FileNotFoundError(f"Arquivo de entrada não encontrado: {input_path}")

    if not input_path.is_file():
        raise IsADirectoryError(f"O caminho de entrada não é um arquivo válido: {input_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)


def _validate_rename_inputs(
    old_columns: Sequence[str],
    new_columns: Sequence[str],
) -> Dict[str, str]:
    """
    Valida as listas de colunas e monta o mapa de renomeação.

    Args:
        old_columns (Sequence[str]): Colunas atuais informadas pelo usuário.
        new_columns (Sequence[str]): Novos nomes informados pelo usuário.

    Returns:
        Dict[str, str]: Mapa entre nome antigo e novo nome.
    """
    normalized_old_columns = [column.strip() for column in old_columns if column and column.strip()]
    normalized_new_columns = [column.strip() for column in new_columns if column and column.strip()]

    if not normalized_old_columns or not normalized_new_columns:
        raise ValueError("As listas de colunas antigas e novas não podem estar vazias.")

    if len(normalized_old_columns) != len(old_columns) or len(normalized_new_columns) != len(new_columns):
        raise ValueError("As listas de colunas não podem conter valores vazios.")

    if len(normalized_old_columns) != len(normalized_new_columns):
        raise ValueError(
            "A quantidade de colunas antigas deve ser igual à quantidade de colunas novas."
        )

    if len(set(normalized_old_columns)) != len(normalized_old_columns):
        raise ValueError("A lista de colunas antigas contém nomes duplicados.")

    if len(set(normalized_new_columns)) != len(normalized_new_columns):
        raise ValueError("A lista de colunas novas contém nomes duplicados.")

    return dict(zip(normalized_old_columns, normalized_new_columns))


def _validate_output_extension(input_path: Path, output_path: Path) -> None:
    """
    Garante que o formato de saída seja compatível com o formato de entrada.

    Args:
        input_path (Path): Caminho do arquivo de entrada.
        output_path (Path): Caminho do arquivo de saída.

    Returns:
        None: Esta função apenas valida a extensão de saída.
    """
    input_extension = input_path.suffix.lower()
    output_extension = output_path.suffix.lower()

    if input_extension not in SUPPORTED_FILE_EXTENSIONS | {".xls"}:
        raise ValueError(f"Extensão de arquivo não suportada: {input_extension}")

    if output_extension != input_extension:
        raise ValueError(
            "O arquivo de saída deve manter a mesma extensão do arquivo de entrada."
        )


def _rename_csv_columns(
    input_path: Path,
    output_path: Path,
    rename_map: Dict[str, str],
    delimiter: Optional[str] = None,
) -> int:
    """
    Renomeia o cabeçalho de um CSV sem carregar o arquivo inteiro em memória.

    Args:
        input_path (Path): Caminho do arquivo CSV de entrada.
        output_path (Path): Caminho do arquivo CSV de saída.
        rename_map (Dict[str, str]): Mapa de renomeação.
        delimiter (Optional[str]): Delimitador informado pelo usuário.

    Returns:
        int: Quantidade de colunas renomeadas.
    """
    csv_settings = _resolve_csv_settings(input_path=input_path, delimiter=delimiter)
    line_ending = _detect_line_ending(input_path=input_path, encoding=csv_settings["encoding"])
    temporary_output_path = _build_temporary_output_path(output_path)

    try:
        with input_path.open("r", encoding=csv_settings["encoding"], newline="") as source_file:
            header_line = source_file.readline()
            if not header_line:
                raise ValueError("O arquivo CSV está vazio e não possui cabeçalho para renomear.")

            current_columns = next(
                csv.reader([header_line], delimiter=csv_settings["delimiter"])
            )
            renamed_columns = _rename_header_columns(
                current_columns=current_columns,
                rename_map=rename_map,
            )

            with temporary_output_path.open(
                "w",
                encoding=csv_settings["encoding"],
                newline="",
            ) as temporary_file:
                temporary_file.write(
                    _serialize_csv_row(
                        columns=renamed_columns,
                        delimiter=csv_settings["delimiter"],
                        line_ending=line_ending,
                    )
                )
                shutil.copyfileobj(source_file, temporary_file, length=1024 * 1024)

        os.replace(temporary_output_path, output_path)
        return sum(
            1
            for current_name, renamed_name in zip(current_columns, renamed_columns)
            if current_name != renamed_name
        )
    finally:
        if temporary_output_path.exists():
            temporary_output_path.unlink(missing_ok=True)


def _resolve_csv_settings(input_path: Path, delimiter: Optional[str]) -> Dict[str, str]:
    """
    Resolve encoding e delimitador de um arquivo CSV.

    Args:
        input_path (Path): Caminho do arquivo CSV.
        delimiter (Optional[str]): Delimitador informado pelo usuário.

    Returns:
        Dict[str, str]: Configurações de leitura do CSV.
    """
    if delimiter:
        detection_result = detect_csv_delimiter(str(input_path))
        detected_encoding = (
            detection_result.get("encoding_used")
            if isinstance(detection_result, dict)
            else None
        )
        return {
            "delimiter": delimiter,
            "encoding": detected_encoding or "utf-8",
        }

    detection_result = detect_csv_delimiter(str(input_path))
    if isinstance(detection_result, dict) and not detection_result.get("error"):
        return {
            "delimiter": detection_result.get("delimiter") or ";",
            "encoding": detection_result.get("encoding_used") or "utf-8",
        }

    return {
        "delimiter": ";",
        "encoding": "utf-8",
    }


def _detect_line_ending(input_path: Path, encoding: str) -> str:
    """
    Detecta a quebra de linha usada no cabeçalho do arquivo CSV.

    Args:
        input_path (Path): Caminho do arquivo CSV.
        encoding (str): Encoding utilizado para abrir o arquivo.

    Returns:
        str: Quebra de linha detectada.
    """
    with input_path.open("r", encoding=encoding, newline="") as source_file:
        header_line = source_file.readline()

    if header_line.endswith("\r\n"):
        return "\r\n"
    if header_line.endswith("\r"):
        return "\r"
    return "\n"


def _serialize_csv_row(columns: Sequence[str], delimiter: str, line_ending: str) -> str:
    """
    Serializa uma linha CSV usando o delimitador informado.

    Args:
        columns (Sequence[str]): Valores da linha.
        delimiter (str): Delimitador do CSV.
        line_ending (str): Quebra de linha a ser usada.

    Returns:
        str: Linha CSV serializada.
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=delimiter, lineterminator=line_ending)
    writer.writerow(columns)
    return buffer.getvalue()


def _rename_xlsx_columns(
    input_path: Path,
    output_path: Path,
    rename_map: Dict[str, str],
    sheet_name: Optional[str] = None,
) -> int:
    """
    Renomeia colunas da primeira linha de uma planilha Excel.

    Args:
        input_path (Path): Caminho do arquivo Excel de entrada.
        output_path (Path): Caminho do arquivo Excel de saída.
        rename_map (Dict[str, str]): Mapa de renomeação.
        sheet_name (Optional[str]): Nome da planilha a ser alterada.

    Returns:
        int: Quantidade de colunas renomeadas.
    """
    workbook = load_workbook(filename=input_path)

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"A planilha '{sheet_name}' não foi encontrada no arquivo Excel informado."
            )
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook[workbook.sheetnames[0]]

    current_columns = [
        "" if cell.value is None else str(cell.value)
        for cell in worksheet[1]
    ]
    if not current_columns:
        raise ValueError("A planilha selecionada não possui cabeçalho para renomear.")

    renamed_columns = _rename_header_columns(
        current_columns=current_columns,
        rename_map=rename_map,
    )

    for index, column_name in enumerate(renamed_columns, start=1):
        worksheet.cell(row=1, column=index, value=column_name)

    workbook.save(output_path)
    return sum(
        1
        for current_name, renamed_name in zip(current_columns, renamed_columns)
        if current_name != renamed_name
    )


def _rename_header_columns(
    current_columns: Sequence[str],
    rename_map: Dict[str, str],
) -> Sequence[str]:
    """
    Renomeia a lista de colunas e valida inconsistências de esquema.

    Args:
        current_columns (Sequence[str]): Cabeçalho atual do arquivo.
        rename_map (Dict[str, str]): Mapa de renomeação.

    Returns:
        Sequence[str]: Novo cabeçalho após a renomeação.
    """
    missing_columns = [
        column_name
        for column_name in rename_map
        if column_name not in current_columns
    ]
    if missing_columns:
        raise ValueError(
            "As seguintes colunas não foram encontradas no arquivo: "
            + ", ".join(missing_columns)
        )

    renamed_columns = [rename_map.get(column_name, column_name) for column_name in current_columns]
    if len(set(renamed_columns)) != len(renamed_columns):
        raise ValueError(
            "A renomeação resultaria em colunas duplicadas no arquivo de saída."
        )

    return renamed_columns


def _build_temporary_output_path(output_path: Path) -> Path:
    """
    Cria um caminho temporário seguro no mesmo diretório do arquivo de saída.

    Args:
        output_path (Path): Caminho final do arquivo de saída.

    Returns:
        Path: Caminho temporário para escrita atômica.
    """
    with tempfile.NamedTemporaryFile(
        dir=output_path.parent,
        prefix=f"{output_path.stem}_",
        suffix=f"{output_path.suffix}.tmp",
        delete=False,
    ) as temporary_file:
        return Path(temporary_file.name)
