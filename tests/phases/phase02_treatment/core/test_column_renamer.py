# -*- coding: utf-8 -*-

import subprocess
import sys

import pandas as pd
from openpyxl import Workbook, load_workbook

from src.phases.phase02_treatment.core.column_renamer import rename_columns_in_file


def test_rename_columns_in_csv_with_detected_delimiter(tmp_path):
    input_file = tmp_path / "input.csv"
    output_file = tmp_path / "output.csv"
    input_file.write_text("old_a,old_b\n1,2\n3,4\n", encoding="utf-8")

    result = rename_columns_in_file(
        input_file=str(input_file),
        old_columns=["old_a", "old_b"],
        new_columns=["new_a", "new_b"],
        output_file=str(output_file),
    )

    assert result["status"] == "success"
    assert result["renamed_count"] == 2

    renamed_dataframe = pd.read_csv(output_file)
    assert renamed_dataframe.columns.tolist() == ["new_a", "new_b"]
    assert renamed_dataframe.iloc[1, 1] == 4


def test_rename_columns_in_xlsx(tmp_path):
    input_file = tmp_path / "input.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Planilha1"
    worksheet.append(["old_a", "old_b"])
    worksheet.append([1, 2])
    workbook.save(input_file)

    result = rename_columns_in_file(
        input_file=str(input_file),
        old_columns=["old_a"],
        new_columns=["new_a"],
        sheet_name="Planilha1",
    )

    assert result["status"] == "success"
    assert result["renamed_count"] == 1

    renamed_workbook = load_workbook(input_file)
    renamed_sheet = renamed_workbook["Planilha1"]
    assert renamed_sheet["A1"].value == "new_a"
    assert renamed_sheet["B1"].value == "old_b"


def test_rename_columns_in_file_rejects_inconsistent_lists(tmp_path):
    input_file = tmp_path / "input.csv"
    input_file.write_text("col_a;col_b\n1;2\n", encoding="utf-8")

    result = rename_columns_in_file(
        input_file=str(input_file),
        old_columns=["col_a", "col_b"],
        new_columns=["novo_a"],
    )

    assert result["status"] == "error"
    assert "quantidade de colunas antigas" in result["message"]


def test_rename_columns_script_wrapper_success(tmp_path):
    input_file = tmp_path / "script_input.csv"
    output_file = tmp_path / "script_output.csv"
    input_file.write_text("name;value\nalpha;10\n", encoding="utf-8")

    command = [
        sys.executable,
        "-m",
        "src.scripts.rename_columns",
        "--input-file",
        str(input_file),
        "--old-columns",
        "name",
        "value",
        "--new-columns",
        "nome",
        "valor",
        "--output-file",
        str(output_file),
        "--delimiter",
        ";",
    ]
    result = subprocess.run(command, capture_output=True, text=True, check=False)

    assert result.returncode == 0
    wrapped_dataframe = pd.read_csv(output_file, sep=";")
    assert wrapped_dataframe.columns.tolist() == ["nome", "valor"]
